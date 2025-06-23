#!/usr/bin/env python3
"""
Excel File Editor MCP Server

このサーバーは、Excel ファイルの読み書きや編集を行うMCPサーバーです。
Model Context Protocolを使用して、Excelファイルの操作を提供します。
"""

import os
import sys
import json
from pathlib import Path
from typing import List, Dict, Any, Optional
from fastmcp import FastMCP
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# MCPアプリケーションの作成
app = FastMCP("Excel File Editor 📊")

@app.tool
def read_excel_file(file_path: str, sheet_name: Optional[str] = None) -> dict:
    """
    Excelファイルを読み取る
    
    Args:
        file_path: Excelファイルのパス
        sheet_name: 読み取るシート名（省略時は最初のシート）
        
    Returns:
        読み取ったデータを含む辞書
    """
    try:
        if not os.path.exists(file_path):
            return {"error": f"ファイルが見つかりません: {file_path}"}
        
        # ファイル拡張子をチェック
        if not file_path.endswith(('.xlsx', '.xls')):
            return {"error": "サポートされていないファイル形式です。.xlsx または .xls ファイルを指定してください。"}
        
        # Excelファイルを読み取る
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_path)
        
        # データを辞書形式に変換し、行番号にヘッダー分（+1）を追加
        data_records = []
        for idx, record in enumerate(df.to_dict('records')):
            record['_excel_row_number'] = idx + 2  # ヘッダーを考慮して+2
            data_records.append(record)
        
        # データを辞書形式に変換
        data = {
            "file_path": file_path,
            "sheet_name": sheet_name or "デフォルトシート",
            "shape": {"rows": len(df), "columns": len(df.columns)},
            "columns": df.columns.tolist(),
            "data": data_records,
            "head": df.head().to_dict('records'),
            "note": "データの行番号は _excel_row_number フィールドで確認できます（Excelの実際の行番号）"
        }
        
        return data
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

@app.tool
def list_excel_sheets(file_path: str) -> dict:
    """
    Excelファイル内のシート一覧を取得する
    
    Args:
        file_path: Excelファイルのパス
        
    Returns:
        シート名のリストを含む辞書
    """
    try:
        if not os.path.exists(file_path):
            return {"error": f"ファイルが見つかりません: {file_path}"}
        
        # Excelファイルを開く
        workbook = openpyxl.load_workbook(file_path)
        sheets = workbook.sheetnames
        
        return {
            "file_path": file_path,
            "sheets": sheets,
            "total_sheets": len(sheets)
        }
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

@app.tool
def write_excel_file(file_path: str, data: List[Dict[str, Any]], sheet_name: str = "Sheet1") -> dict:
    """
    データをExcelファイルに書き込む
    
    Args:
        file_path: 保存するExcelファイルのパス
        data: 書き込むデータ（辞書のリスト）
        sheet_name: シート名
        
    Returns:
        処理結果を含む辞書
    """
    try:
        # データをDataFrameに変換
        df = pd.DataFrame(data)
        
        # Excelファイルに書き込む
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return {
            "status": "success",
            "file_path": file_path,
            "sheet_name": sheet_name,
            "rows_written": len(data),
            "columns": list(df.columns) if not df.empty else []
        }
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

@app.tool
def update_excel_cell(file_path: str, sheet_name: str, row: int, column: str, value: Any) -> dict:
    """
    Excelファイルの特定のセルを更新する
    
    Args:
        file_path: Excelファイルのパス
        sheet_name: シート名
        row: 行番号（1から開始）
        column: 列名（A, B, C...）
        value: 設定する値
        
    Returns:
        処理結果を含む辞書
    """
    try:
        if not os.path.exists(file_path):
            return {"error": f"ファイルが見つかりません: {file_path}"}
        
        # Excelファイルを開く
        workbook = openpyxl.load_workbook(file_path)
        
        # シートを取得
        if sheet_name not in workbook.sheetnames:
            return {"error": f"シート '{sheet_name}' が見つかりません"}
        
        worksheet = workbook[sheet_name]
        
        # セルを更新
        cell_address = f"{column}{row}"
        worksheet[cell_address] = value
        
        # ファイルを保存
        workbook.save(file_path)
        
        return {
            "status": "success",
            "file_path": file_path,
            "sheet_name": sheet_name,
            "cell": cell_address,
            "value": value
        }
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

@app.tool
def format_excel_cells(file_path: str, sheet_name: str, cell_range: str, 
                      font_color: Optional[str] = None, bg_color: Optional[str] = None,
                      bold: bool = False, font_size: Optional[int] = None) -> dict:
    """
    Excelファイルのセル範囲をフォーマットする
    
    Args:
        file_path: Excelファイルのパス
        sheet_name: シート名
        cell_range: セル範囲（例: "A1:C3"）
        font_color: フォント色（16進数カラーコード、例: "FF0000"）
        bg_color: 背景色（16進数カラーコード、例: "FFFF00"）
        bold: 太字にするかどうか
        font_size: フォントサイズ
        
    Returns:
        処理結果を含む辞書
    """
    try:
        if not os.path.exists(file_path):
            return {"error": f"ファイルが見つかりません: {file_path}"}
        
        # Excelファイルを開く
        workbook = openpyxl.load_workbook(file_path)
        
        # シートを取得
        if sheet_name not in workbook.sheetnames:
            return {"error": f"シート '{sheet_name}' が見つかりません"}
        
        worksheet = workbook[sheet_name]
        
        # フォーマットを適用
        for row in worksheet[cell_range]:
            for cell in row:
                # フォントの設定
                font_kwargs = {}
                if font_color:
                    font_kwargs['color'] = font_color
                if bold:
                    font_kwargs['bold'] = True
                if font_size:
                    font_kwargs['size'] = font_size
                
                if font_kwargs:
                    cell.font = Font(**font_kwargs)
                
                # 背景色の設定
                if bg_color:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        # ファイルを保存
        workbook.save(file_path)
        
        return {
            "status": "success",
            "file_path": file_path,
            "sheet_name": sheet_name,
            "cell_range": cell_range,
            "formatting_applied": True
        }
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

@app.tool
def add_excel_sheet(file_path: str, sheet_name: str, data: Optional[List[Dict[str, Any]]] = None) -> dict:
    """
    Excelファイルに新しいシートを追加する
    
    Args:
        file_path: Excelファイルのパス
        sheet_name: 新しいシート名
        data: 追加するデータ（省略可能）
        
    Returns:
        処理結果を含む辞書
    """
    try:
        # ファイルが存在しない場合は新規作成
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
        else:
            workbook = openpyxl.load_workbook(file_path)
        
        # シートを追加
        if sheet_name in workbook.sheetnames:
            return {"error": f"シート '{sheet_name}' は既に存在します"}
        
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # データが指定されている場合は書き込む
        if data:
            df = pd.DataFrame(data)
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)
        
        # ファイルを保存
        workbook.save(file_path)
        
        return {
            "status": "success",
            "file_path": file_path,
            "sheet_name": sheet_name,
            "rows_added": len(data) if data else 0
        }
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

@app.tool
def delete_excel_sheet(file_path: str, sheet_name: str) -> dict:
    """
    Excelファイルからシートを削除する
    
    Args:
        file_path: Excelファイルのパス
        sheet_name: 削除するシート名
        
    Returns:
        処理結果を含む辞書
    """
    try:
        if not os.path.exists(file_path):
            return {"error": f"ファイルが見つかりません: {file_path}"}
        
        # Excelファイルを開く
        workbook = openpyxl.load_workbook(file_path)
        
        # シートを削除
        if sheet_name not in workbook.sheetnames:
            return {"error": f"シート '{sheet_name}' が見つかりません"}
        
        if len(workbook.sheetnames) == 1:
            return {"error": "最後のシートは削除できません"}
        
        workbook.remove(workbook[sheet_name])
        
        # ファイルを保存
        workbook.save(file_path)
        
        return {
            "status": "success",
            "file_path": file_path,
            "deleted_sheet": sheet_name,
            "remaining_sheets": workbook.sheetnames
        }
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

@app.tool
def excel_data_summary(file_path: str, sheet_name: Optional[str] = None) -> dict:
    """
    Excelファイルのデータ概要を取得する
    
    Args:
        file_path: Excelファイルのパス
        sheet_name: 分析するシート名（省略時は最初のシート）
        
    Returns:
        データ概要を含む辞書
    """
    try:
        if not os.path.exists(file_path):
            return {"error": f"ファイルが見つかりません: {file_path}"}
        
        # Excelファイルを読み取る
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_path)
        
        # データ概要を生成
        summary = {
            "file_path": file_path,
            "sheet_name": sheet_name or "デフォルトシート",
            "shape": {"rows": len(df), "columns": len(df.columns)},
            "columns": df.columns.tolist(),
            "data_types": df.dtypes.astype(str).to_dict(),
            "null_counts": df.isnull().sum().to_dict(),
            "memory_usage": df.memory_usage(deep=True).sum(),
        }
        
        # 数値列の統計情報
        numeric_columns = df.select_dtypes(include=['number']).columns
        if len(numeric_columns) > 0:
            summary["numeric_summary"] = df[numeric_columns].describe().to_dict()
        
        return summary
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

@app.tool
def find_excel_row_by_content(file_path: str, sheet_name: str, search_column: str, search_value: str) -> dict:
    """
    指定した列の値で行を検索し、Excelの実際の行番号を取得する
    
    Args:
        file_path: Excelファイルのパス
        sheet_name: シート名
        search_column: 検索する列名
        search_value: 検索する値
        
    Returns:
        検索結果を含む辞書
    """
    try:
        if not os.path.exists(file_path):
            return {"error": f"ファイルが見つかりません: {file_path}"}
        
        # Excelファイルを読み取る
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        # 指定した値を検索
        matches = df[df[search_column].astype(str).str.contains(search_value, na=False, case=False)]
        
        if matches.empty:
            return {
                "found": False,
                "message": f"'{search_value}' が列 '{search_column}' で見つかりませんでした"
            }
        
        # マッチした行の情報を返す
        results = []
        for idx, row in matches.iterrows():
            excel_row_number = idx + 2  # ヘッダーを考慮して+2
            results.append({
                "excel_row_number": excel_row_number,
                "pandas_index": idx,
                "data": row.to_dict()
            })
        
        return {
            "found": True,
            "search_column": search_column,
            "search_value": search_value,
            "matches": results,
            "total_matches": len(results)
        }
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

def main():
    app.run()

if __name__ == "__main__":
    main()
