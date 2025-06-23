#!/usr/bin/env python3
"""
Excel操作機能を直接テストする
"""

import os
import json
import pandas as pd
import openpyxl
from typing import List, Dict, Any, Optional
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def read_excel_file(file_path: str, sheet_name: Optional[str] = None) -> dict:
    """Excelファイルを読み取る"""
    try:
        if not os.path.exists(file_path):
            return {"error": f"ファイルが見つかりません: {file_path}"}
        
        if not file_path.endswith(('.xlsx', '.xls')):
            return {"error": "サポートされていないファイル形式です。"}
        
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_path)
        
        data = {
            "file_path": file_path,
            "sheet_name": sheet_name or "デフォルトシート",
            "shape": {"rows": len(df), "columns": len(df.columns)},
            "columns": df.columns.tolist(),
            "data": df.to_dict('records'),
            "head": df.head().to_dict('records')
        }
        
        return data
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

def list_excel_sheets(file_path: str) -> dict:
    """Excelファイル内のシート一覧を取得する"""
    try:
        if not os.path.exists(file_path):
            return {"error": f"ファイルが見つかりません: {file_path}"}
        
        workbook = openpyxl.load_workbook(file_path)
        sheets = workbook.sheetnames
        
        return {
            "file_path": file_path,
            "sheets": sheets,
            "total_sheets": len(sheets)
        }
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

def write_excel_file(file_path: str, data: List[Dict[str, Any]], sheet_name: str = "Sheet1") -> dict:
    """データをExcelファイルに書き込む"""
    try:
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return {
            "status": "success",
            "file_path": file_path,
            "sheet_name": sheet_name,
            "rows_written": len(data),
            "columns": list(df.columns) if not df.empty else []
        }
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

def update_excel_cell(file_path: str, sheet_name: str, row: int, column: str, value: Any) -> dict:
    """Excelファイルの特定のセルを更新する"""
    try:
        if not os.path.exists(file_path):
            return {"error": f"ファイルが見つかりません: {file_path}"}
        
        workbook = openpyxl.load_workbook(file_path)
        
        if sheet_name not in workbook.sheetnames:
            return {"error": f"シート '{sheet_name}' が見つかりません"}
        
        worksheet = workbook[sheet_name]
        cell_address = f"{column}{row}"
        worksheet[cell_address] = value
        
        workbook.save(file_path)
        
        return {
            "status": "success",
            "file_path": file_path,
            "sheet_name": sheet_name,
            "cell": cell_address,
            "value": value
        }
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

def add_excel_sheet(file_path: str, sheet_name: str, data: Optional[List[Dict[str, Any]]] = None) -> dict:
    """Excelファイルに新しいシートを追加する"""
    try:
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
        else:
            workbook = openpyxl.load_workbook(file_path)
        
        if sheet_name in workbook.sheetnames:
            return {"error": f"シート '{sheet_name}' は既に存在します"}
        
        worksheet = workbook.create_sheet(title=sheet_name)
        
        if data:
            df = pd.DataFrame(data)
            for r in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(r)
        
        workbook.save(file_path)
        
        return {
            "status": "success",
            "file_path": file_path,
            "sheet_name": sheet_name,
            "rows_added": len(data) if data else 0
        }
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

def excel_data_summary(file_path: str, sheet_name: Optional[str] = None) -> dict:
    """Excelファイルのデータ概要を取得する"""
    try:
        if not os.path.exists(file_path):
            return {"error": f"ファイルが見つかりません: {file_path}"}
        
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_path)
        
        summary = {
            "file_path": file_path,
            "sheet_name": sheet_name or "デフォルトシート",
            "shape": {"rows": len(df), "columns": len(df.columns)},
            "columns": df.columns.tolist(),
            "data_types": df.dtypes.astype(str).to_dict(),
            "null_counts": df.isnull().sum().to_dict(),
            "memory_usage": df.memory_usage(deep=True).sum(),
        }
        
        # 数値列の統計情報
        numeric_columns = df.select_dtypes(include=['number']).columns
        if len(numeric_columns) > 0:
            summary["numeric_summary"] = df[numeric_columns].describe().to_dict()
        
        return summary
    except Exception as e:
        return {"error": f"エラーが発生しました: {str(e)}"}

def test_all_functions():
    """全ての機能をテスト"""
    
    sample_file = "sample_data.xlsx"
    
    if not os.path.exists(sample_file):
        print("❌ サンプルファイルが見つかりません。create_sample.py を実行してください。")
        return
    
    print("🧪 Excel機能直接テスト開始")
    print("=" * 60)
    
    # テスト1: ファイル読み取り
    print("\n1. 📖 ファイル読み取りテスト")
    result = read_excel_file(sample_file, "従業員データ")
    if "error" not in result:
        print(f"✅ 成功: {result['shape']['rows']}行, {result['shape']['columns']}列")
        print(f"   列: {result['columns']}")
        print(f"   データサンプル: {result['head'][0] if result['head'] else 'なし'}")
    else:
        print(f"❌ エラー: {result['error']}")
    
    # テスト2: シート一覧
    print("\n2. 📄 シート一覧テスト")
    result = list_excel_sheets(sample_file)
    if "error" not in result:
        print(f"✅ 成功: {result['total_sheets']}シート")
        print(f"   シート名: {result['sheets']}")
    else:
        print(f"❌ エラー: {result['error']}")
    
    # テスト3: データ概要
    print("\n3. 📊 データ概要テスト")
    result = excel_data_summary(sample_file, "売上データ")
    if "error" not in result:
        print(f"✅ 成功: {result['shape']}")
        print(f"   データ型: {result['data_types']}")
        print(f"   欠損値: {result['null_counts']}")
        if "numeric_summary" in result:
            print(f"   数値列の統計: あり ({len(result['numeric_summary'])}列)")
    else:
        print(f"❌ エラー: {result['error']}")
    
    # テスト4: セル更新
    print("\n4. ✏️ セル更新テスト")
    import shutil
    backup_file = "sample_data_backup.xlsx"
    shutil.copy2(sample_file, backup_file)
    
    result = update_excel_cell(sample_file, "従業員データ", 1, "F", "更新テスト")
    if "error" not in result:
        print(f"✅ 成功: セル{result['cell']}を更新")
        print(f"   値: {result['value']}")
    else:
        print(f"❌ エラー: {result['error']}")
    
    # バックアップから復元
    shutil.copy2(backup_file, sample_file)
    os.remove(backup_file)
    
    # テスト5: 新しいシート追加
    print("\n5. ➕ シート追加テスト")
    test_data = [
        {"項目": "テスト1", "値": 100},
        {"項目": "テスト2", "値": 200}
    ]
    
    backup_file = "sample_data_backup.xlsx"
    shutil.copy2(sample_file, backup_file)
    
    result = add_excel_sheet(sample_file, "テストシート", test_data)
    if "error" not in result:
        print(f"✅ 成功: シート'{result['sheet_name']}'を追加")
        print(f"   行数: {result['rows_added']}")
        
        # 確認
        sheets_result = list_excel_sheets(sample_file)
        print(f"   現在のシート: {sheets_result['sheets']}")
    else:
        print(f"❌ エラー: {result['error']}")
    
    # バックアップから復元
    shutil.copy2(backup_file, sample_file)
    os.remove(backup_file)
    
    # テスト6: 新しいファイル作成
    print("\n6. 📝 新ファイル作成テスト")
    new_file = "test_output.xlsx"
    test_data = [
        {"名前": "新規太郎", "年齢": 25, "職業": "テスター"},
        {"名前": "作成花子", "年齢": 30, "職業": "エンジニア"}
    ]
    
    result = write_excel_file(new_file, test_data, "新規データ")
    if "error" not in result:
        print(f"✅ 成功: ファイル'{result['file_path']}'を作成")
        print(f"   行数: {result['rows_written']}")
        print(f"   列: {result['columns']}")
        
        # 作成されたファイルを確認
        verify_result = read_excel_file(new_file, "新規データ")
        if "error" not in verify_result:
            print(f"   検証: データ正常読み込み ({verify_result['shape']})")
        
        # ファイルを削除
        if os.path.exists(new_file):
            os.remove(new_file)
            print(f"   🗑️ テストファイルを削除")
    else:
        print(f"❌ エラー: {result['error']}")
    
    print("\n" + "=" * 60)
    print("🎉 全テスト完了!")

if __name__ == "__main__":
    test_all_functions()
