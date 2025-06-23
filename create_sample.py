#!/usr/bin/env python3
"""
サンプルExcelファイルを作成してMCPサーバーをテストする
"""

import pandas as pd
import os

def create_sample_excel():
    """サンプルExcelファイルを作成"""
    
    # サンプルデータ
    employees_data = [
        {"従業員ID": "E001", "名前": "田中太郎", "部署": "開発部", "年齢": 30, "給与": 5000000},
        {"従業員ID": "E002", "名前": "佐藤花子", "部署": "営業部", "年齢": 28, "給与": 4500000},
        {"従業員ID": "E003", "名前": "山田次郎", "部署": "開発部", "年齢": 35, "給与": 6000000},
        {"従業員ID": "E004", "名前": "鈴木美咲", "部署": "人事部", "年齢": 32, "給与": 4800000},
        {"従業員ID": "E005", "名前": "高橋一郎", "部署": "営業部", "年齢": 29, "給与": 4700000}
    ]
    
    sales_data = [
        {"月": "2024-01", "売上": 1200000, "利益": 240000},
        {"月": "2024-02", "売上": 1350000, "利益": 270000},
        {"月": "2024-03", "売上": 1100000, "利益": 220000},
        {"月": "2024-04", "売上": 1450000, "利益": 290000},
        {"月": "2024-05", "売上": 1600000, "利益": 320000}
    ]
    
    # Excelファイル作成
    file_path = "sample_data.xlsx"
    
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # 従業員データシート
        df_employees = pd.DataFrame(employees_data)
        df_employees.to_excel(writer, sheet_name='従業員データ', index=False)
        
        # 売上データシート
        df_sales = pd.DataFrame(sales_data)
        df_sales.to_excel(writer, sheet_name='売上データ', index=False)
    
    print(f"✅ サンプルExcelファイルを作成しました: {file_path}")
    return file_path

if __name__ == "__main__":
    sample_file = create_sample_excel()
    
    # ファイル情報を表示
    if os.path.exists(sample_file):
        file_size = os.path.getsize(sample_file)
        print(f"📊 ファイルサイズ: {file_size} bytes")
        
        # シート確認
        import openpyxl
        wb = openpyxl.load_workbook(sample_file)
        print(f"📄 シート一覧: {wb.sheetnames}")
        
        # データ確認
        for sheet_name in wb.sheetnames:
            df = pd.read_excel(sample_file, sheet_name=sheet_name)
            print(f"📋 '{sheet_name}' シート: {df.shape[0]}行, {df.shape[1]}列")
            print(f"   列: {list(df.columns)}")
            print()
