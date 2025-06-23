#!/usr/bin/env python3
"""
Test script for Excel MCP Server
"""

import json
import tempfile
import os
from pathlib import Path
import pandas as pd

def test_excel_operations():
    """Excel操作のテスト"""
    
    # テスト用データ
    test_data = [
        {"名前": "田中太郎", "年齢": 30, "職業": "エンジニア"},
        {"名前": "佐藤花子", "年齢": 25, "職業": "デザイナー"},
        {"名前": "山田次郎", "年齢": 35, "職業": "営業"}
    ]
    
    # 一時ファイルを作成
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        temp_path = tmp_file.name
    
    try:
        print("Excel MCP Server テスト開始")
        print("=" * 50)
        
        # 1. Excelファイル作成テスト
        print("1. Excelファイル作成テスト")
        df = pd.DataFrame(test_data)
        df.to_excel(temp_path, index=False, sheet_name="テストデータ")
        print(f"✓ ファイル作成成功: {temp_path}")
        
        # 2. ファイル読み取りテスト
        print("\n2. ファイル読み取りテスト")
        read_df = pd.read_excel(temp_path)
        print(f"✓ データ読み取り成功: {len(read_df)}行, {len(read_df.columns)}列")
        print(f"  列: {list(read_df.columns)}")
        
        # 3. シート一覧テスト
        print("\n3. シート一覧テスト")
        import openpyxl
        wb = openpyxl.load_workbook(temp_path)
        sheets = wb.sheetnames
        print(f"✓ シート一覧: {sheets}")
        
        # 4. セル更新テスト
        print("\n4. セル更新テスト")
        ws = wb.active
        ws['D1'] = "追加列"
        ws['D2'] = "追加データ1"
        wb.save(temp_path)
        print("✓ セル更新成功")
        
        # 5. データ概要テスト
        print("\n5. データ概要テスト")
        updated_df = pd.read_excel(temp_path)
        print(f"✓ 更新後のデータ: {updated_df.shape}")
        print(f"  データ型: {dict(updated_df.dtypes)}")
        print(f"  欠損値: {dict(updated_df.isnull().sum())}")
        
        print("\n" + "=" * 50)
        print("✅ 全テスト完了")
        
    except Exception as e:
        print(f"❌ テストエラー: {e}")
    finally:
        # 一時ファイルを削除
        if os.path.exists(temp_path):
            os.unlink(temp_path)
            print(f"🗑️ 一時ファイル削除: {temp_path}")

def test_mcp_tools():
    """MCPツールの模擬テスト"""
    print("\nMCPツール模擬テスト")
    print("=" * 30)
    
    # 模擬的なツール呼び出し例
    tools_examples = [
        {
            "tool": "read_excel_file",
            "description": "Excelファイルの読み取り",
            "example_args": {
                "file_path": "sample.xlsx",
                "sheet_name": "Sheet1"
            }
        },
        {
            "tool": "write_excel_file", 
            "description": "Excelファイルの書き込み",
            "example_args": {
                "file_path": "output.xlsx",
                "data": [{"col1": "value1", "col2": "value2"}],
                "sheet_name": "NewSheet"
            }
        },
        {
            "tool": "update_excel_cell",
            "description": "セルの更新",
            "example_args": {
                "file_path": "sample.xlsx",
                "sheet_name": "Sheet1", 
                "row": 1,
                "column": "A",
                "value": "新しい値"
            }
        },
        {
            "tool": "format_excel_cells",
            "description": "セルのフォーマット",
            "example_args": {
                "file_path": "sample.xlsx",
                "sheet_name": "Sheet1",
                "cell_range": "A1:C3",
                "font_color": "FF0000",
                "bold": True
            }
        }
    ]
    
    for tool in tools_examples:
        print(f"📋 {tool['tool']}: {tool['description']}")
        print(f"   例: {json.dumps(tool['example_args'], ensure_ascii=False, indent=6)}")
        print()

if __name__ == "__main__":
    test_excel_operations()
    test_mcp_tools()
